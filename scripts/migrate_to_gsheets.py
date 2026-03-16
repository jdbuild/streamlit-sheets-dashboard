from __future__ import annotations

import argparse
import colorsys
from datetime import date, datetime, time, timedelta
from pathlib import Path
import re
from xml.etree import ElementTree as ET

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils.cell import range_boundaries

from planning.config import get_config


def _sheets_service(credentials: Credentials):
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def _drive_service(credentials: Credentials):
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _load_credentials(path: Path) -> Credentials:
    return Credentials.from_authorized_user_file(
        path,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )


def migrate(local_xlsm: Path, credentials_path: Path, title: str) -> str:
    credentials = _load_credentials(credentials_path)
    workbook = load_workbook(local_xlsm, data_only=False, keep_vba=True)
    if not workbook.worksheets:
        raise ValueError(f"Workbook contains no worksheets: {local_xlsm}")
    theme_colors = _theme_colors(workbook)
    spreadsheet = _sheets_service(credentials).spreadsheets().create(
        body={"properties": {"title": title, "locale": "de_DE"}}
    ).execute()
    spreadsheet_id = spreadsheet["spreadsheetId"]
    sheet_metadata = _sheets_service(credentials).spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    default_sheet_id = sheet_metadata["sheets"][0]["properties"]["sheetId"]
    requests = [
        {
            "updateSheetProperties": {
                "properties": {"sheetId": default_sheet_id, "title": workbook.worksheets[0].title},
                "fields": "title",
            }
        }
    ]
    for sheet in workbook.worksheets[1:]:
        requests.append({"addSheet": {"properties": {"title": sheet.title}}})
    _sheets_service(credentials).spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": requests},
    ).execute()
    metadata = _sheets_service(credentials).spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    title_to_id = {item["properties"]["title"]: item["properties"]["sheetId"] for item in metadata["sheets"]}
    for sheet in workbook.worksheets:
        row_count = _required_row_count(sheet)
        column_count = _required_column_count(sheet)
        sizing_requests = [
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": title_to_id[sheet.title],
                        "gridProperties": {
                            "rowCount": row_count,
                            "columnCount": column_count,
                        },
                    },
                    "fields": "gridProperties.rowCount,gridProperties.columnCount",
                }
            }
        ]
        values = []
        for row in sheet.iter_rows():
            values.append([_serialize_cell_value(cell.value) for cell in row])
        _sheets_service(credentials).spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": sizing_requests},
        ).execute()
        _sheets_service(credentials).spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet.title}'!A1",
            valueInputOption="USER_ENTERED",
            body={"values": values},
        ).execute()
        requests = []
        for merged_range in sheet.merged_cells.ranges:
            requests.append(
                {
                    "mergeCells": {
                        "range": _a1_to_grid_range(title_to_id[sheet.title], str(merged_range)),
                        "mergeType": "MERGE_ALL",
                    }
                }
            )
        for col_letter, dim in sheet.column_dimensions.items():
            if dim.width:
                start_index = _col_to_index(col_letter)
                requests.append(
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": title_to_id[sheet.title],
                                "dimension": "COLUMNS",
                                "startIndex": start_index,
                                "endIndex": start_index + 1,
                            },
                            "properties": {"pixelSize": int(dim.width * 7)},
                            "fields": "pixelSize",
                        }
                    }
                )
        for row_idx, dim in sheet.row_dimensions.items():
            if dim.height:
                requests.append(
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": title_to_id[sheet.title],
                                "dimension": "ROWS",
                                "startIndex": row_idx - 1,
                                "endIndex": row_idx,
                            },
                            "properties": {"pixelSize": int(dim.height * 1.33)},
                            "fields": "pixelSize",
                        }
                    }
                )
        if sheet.title == "P-XYTemplate":
            requests.extend(_background_fill_requests(sheet, title_to_id[sheet.title], theme_colors))
        if requests:
            for batch in _chunked(requests, 200):
                _sheets_service(credentials).spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={"requests": batch},
                ).execute()
    return spreadsheet_id


def _col_to_index(col_letter: str) -> int:
    index = 0
    for char in col_letter:
        index = index * 26 + (ord(char.upper()) - 64)
    return index - 1


def _required_row_count(sheet) -> int:
    max_row = max(sheet.max_row, 1)
    for row_idx in sheet.row_dimensions:
        if isinstance(row_idx, int):
            max_row = max(max_row, row_idx)
    for merged_range in sheet.merged_cells.ranges:
        _, _, _, end_row = range_boundaries(str(merged_range))
        max_row = max(max_row, end_row)
    return max_row


def _required_column_count(sheet) -> int:
    max_column = max(sheet.max_column, 1)
    for col_letter in sheet.column_dimensions:
        max_column = max(max_column, _col_to_index(col_letter) + 1)
    for merged_range in sheet.merged_cells.ranges:
        _, _, end_col, _ = range_boundaries(str(merged_range))
        max_column = max(max_column, end_col)
    return max_column


def _serialize_cell_value(value):
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, str):
        return _normalize_formula(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    if hasattr(value, "text"):
        return _normalize_formula(value.text)
    return _normalize_formula(str(value))


def _normalize_formula(value: str) -> str:
    if not value.startswith(("=", "{=")):
        return value
    chars: list[str] = []
    in_string = False
    for idx, char in enumerate(value):
        if char == '"':
            in_string = not in_string
            chars.append(char)
            continue
        if char == "." and not in_string:
            prev_char = value[idx - 1] if idx > 0 else ""
            next_char = value[idx + 1] if idx + 1 < len(value) else ""
            if prev_char.isdigit() and next_char.isdigit():
                chars.append(",")
                continue
        if char == "," and not in_string:
            chars.append(";")
            continue
        chars.append(char)
    return "".join(chars)


def _background_fill_requests(sheet, sheet_id: int, theme_colors: list[str]) -> list[dict]:
    requests: list[dict] = []
    for row in sheet.iter_rows():
        for cell in row:
            color = _google_fill_color(cell.fill, theme_colors)
            if color is None:
                continue
            requests.append(
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": cell.row - 1,
                            "endRowIndex": cell.row,
                            "startColumnIndex": cell.column - 1,
                            "endColumnIndex": cell.column,
                        },
                        "cell": {"userEnteredFormat": {"backgroundColor": color}},
                        "fields": "userEnteredFormat.backgroundColor",
                    }
                }
            )
    return requests


def _google_fill_color(fill, theme_colors: list[str]) -> dict | None:
    if getattr(fill, "fill_type", None) != "solid":
        return None
    fg_color = getattr(fill, "fgColor", None)
    rgb = _resolve_openpyxl_color(fg_color, theme_colors)
    if not rgb:
        return None
    if rgb in {"000000", "FFFFFF"}:
        return None
    return {
        "red": int(rgb[0:2], 16) / 255,
        "green": int(rgb[2:4], 16) / 255,
        "blue": int(rgb[4:6], 16) / 255,
    }


def _chunked(items: list[dict], size: int) -> list[list[dict]]:
    return [items[idx : idx + size] for idx in range(0, len(items), size)]


def _theme_colors(workbook) -> list[str]:
    if not workbook.loaded_theme:
        return []
    root = ET.fromstring(workbook.loaded_theme)
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    clr_scheme = root.find(".//a:clrScheme", ns)
    if clr_scheme is None:
        return []
    scheme_colors: dict[str, str] = {}
    for child in list(clr_scheme):
        color_node = list(child)[0] if list(child) else None
        if color_node is None:
            continue
        value = color_node.attrib.get("lastClr") or color_node.attrib.get("val")
        if not value:
            continue
        match = re.search(r"([0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$", value)
        if match:
            scheme_colors[child.tag.rsplit("}", 1)[-1]] = match.group(1)[-6:].upper()
    # Excel cell style theme indexes use this logical order rather than raw clrScheme order.
    theme_order = [
        "lt1",
        "dk1",
        "lt2",
        "dk2",
        "accent1",
        "accent2",
        "accent3",
        "accent4",
        "accent5",
        "accent6",
        "hlink",
        "folHlink",
    ]
    return [scheme_colors[name] for name in theme_order if name in scheme_colors]


def _resolve_openpyxl_color(color, theme_colors: list[str]) -> str | None:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    tint = float(getattr(color, "tint", 0) or 0)
    base_rgb: str | None = None
    if color_type == "rgb":
        match = re.search(r"([0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$", str(getattr(color, "rgb", "")))
        if match:
            base_rgb = match.group(1)[-6:].upper()
    elif color_type == "theme":
        theme_index = getattr(color, "theme", None)
        if isinstance(theme_index, int) and 0 <= theme_index < len(theme_colors):
            base_rgb = theme_colors[theme_index]
    elif color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if isinstance(indexed, int) and 0 <= indexed < len(COLOR_INDEX):
            base_rgb = COLOR_INDEX[indexed][-6:].upper()
    if base_rgb is None:
        return None
    return _apply_tint(base_rgb, tint)


def _apply_tint(rgb: str, tint: float) -> str:
    if tint == 0:
        return rgb
    red = int(rgb[0:2], 16) / 255
    green = int(rgb[2:4], 16) / 255
    blue = int(rgb[4:6], 16) / 255
    hue, lightness, saturation = colorsys.rgb_to_hls(red, green, blue)
    if tint < 0:
        lightness *= 1 + tint
    else:
        lightness += (1 - lightness) * tint
    red, green, blue = colorsys.hls_to_rgb(hue, max(0, min(1, lightness)), saturation)
    return f"{round(red * 255):02X}{round(green * 255):02X}{round(blue * 255):02X}"


def _a1_to_grid_range(sheet_id: int, a1_range: str) -> dict:
    start, end = a1_range.split(":")
    start_col = "".join(filter(str.isalpha, start))
    start_row = int("".join(filter(str.isdigit, start)))
    end_col = "".join(filter(str.isalpha, end))
    end_row = int("".join(filter(str.isdigit, end)))
    return {
        "sheetId": sheet_id,
        "startRowIndex": start_row - 1,
        "endRowIndex": end_row,
        "startColumnIndex": _col_to_index(start_col),
        "endColumnIndex": _col_to_index(end_col) + 1,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default="data/sandbox.xlsm")
    parser.add_argument("--credentials", required=True)
    parser.add_argument("--title", default="Planning Template")
    args = parser.parse_args()
    spreadsheet_id = migrate(Path(args.source), Path(args.credentials), args.title)
    print(spreadsheet_id)


if __name__ == "__main__":
    main()
