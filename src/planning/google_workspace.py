from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import json
import re

import gspread
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook

from planning.models import TEMPLATE_SHEET


SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


@dataclass(slots=True)
class GoogleWorkspaceClient:
    credentials_payload: dict

    @classmethod
    def from_authorized_user_file(cls, path: str | Path) -> "GoogleWorkspaceClient":
        credentials_path = Path(path)
        return cls(json.loads(credentials_path.read_text()))

    def _credentials(self) -> Credentials:
        creds = Credentials.from_authorized_user_info(self.credentials_payload, scopes=SCOPES)
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            self.credentials_payload.update(json.loads(creds.to_json()))
        return creds

    def gspread_client(self) -> gspread.Client:
        return gspread.authorize(self._credentials())

    def drive_service(self):
        return build("drive", "v3", credentials=self._credentials(), cache_discovery=False)

    def sheets_service(self):
        return build("sheets", "v4", credentials=self._credentials(), cache_discovery=False)

    def open_workspace(self, sheet_id: str):
        return self.gspread_client().open_by_key(sheet_id)

    def copy_spreadsheet(self, template_sheet_id: str, title: str) -> str:
        copied = (
            self.drive_service()
            .files()
            .copy(fileId=template_sheet_id, body={"name": title})
            .execute()
        )
        return copied["id"]

    def duplicate_project_template(self, spreadsheet_id: str, new_title: str) -> None:
        spreadsheet = self.open_workspace(spreadsheet_id)
        template_worksheet = spreadsheet.worksheet(TEMPLATE_SHEET)
        requests = [
            {
                "duplicateSheet": {
                    "sourceSheetId": template_worksheet.id,
                    "newSheetName": new_title,
                }
            }
        ]
        self.sheets_service().spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body={"requests": requests}
        ).execute()

    def spreadsheet_to_workbook(self, spreadsheet_id: str):
        export_bytes = (
            self.drive_service()
            .files()
            .export(fileId=spreadsheet_id, mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            .execute()
        )
        return load_workbook(BytesIO(export_bytes), data_only=True)


def slugify_project_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", name.strip())
    return cleaned.strip("-") or "new-project"
def load_workbook_from_local_file(path: str | Path):
    return load_workbook(Path(path), data_only=False, keep_vba=True)
