from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from planning.models import (
    BLOCK_START_ROWS,
    MONTH_END_COL,
    MONTH_HEADER_ROW,
    MONTH_START_COL,
    PERSON_BLOCK_HEIGHT,
    TEMPLATE_SHEET,
    USER_RANGE,
    WP_RANGE,
    UserDimensionRow,
    WPDimensionRow,
)


@dataclass(slots=True)
class ParsedProjectBlock:
    sheet_title: str
    block_start_row: int
    source_row: int
    person_slot: str | None
    userid: str | None
    person_name: str | None
    role: str | None
    optional_attr: str | None
    wp_label_raw: str | None
    annual_summary: dict[str, float | None]
    monthly_values: list[tuple[date, str, float | None]]
    summary_monthly_values: list[tuple[date, str, float | None]]


@dataclass(slots=True)
class ParsedWorkbook:
    project_sheet_names: list[str]
    user_dimensions: list[UserDimensionRow]
    wp_dimensions: list[WPDimensionRow]
    project_blocks: list[ParsedProjectBlock]


def _excel_date_to_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    raise ValueError(f"Unsupported Excel date value: {value!r}")


def _looks_like_excel_date(value: object) -> bool:
    return isinstance(value, (datetime, date, int, float))


def _to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_workbook(path: Path):
    return load_workbook(path, data_only=True, keep_vba=True)


def parse_local_xlsm(path: str | Path) -> ParsedWorkbook:
    workbook = _load_workbook(Path(path))
    return parse_openpyxl_workbook(workbook)


def parse_openpyxl_workbook(workbook) -> ParsedWorkbook:
    allowed_project_sheets = _budget_project_sheet_names(workbook)
    project_sheet_names = [ws.title for ws in workbook.worksheets if ws.title.startswith("P-")]
    if allowed_project_sheets:
        project_sheet_names = [sheet_name for sheet_name in project_sheet_names if sheet_name in allowed_project_sheets]
    dimension_sheet = workbook[TEMPLATE_SHEET] if TEMPLATE_SHEET in workbook.sheetnames else workbook[project_sheet_names[0]]
    users = parse_user_dimensions(dimension_sheet)
    wps = parse_wp_dimensions(dimension_sheet)
    project_blocks: list[ParsedProjectBlock] = []
    for sheet_name in project_sheet_names:
        if sheet_name == TEMPLATE_SHEET:
            continue
        project_blocks.extend(parse_project_sheet(workbook[sheet_name]))
    return ParsedWorkbook(
        project_sheet_names=project_sheet_names,
        user_dimensions=users,
        wp_dimensions=wps,
        project_blocks=project_blocks,
    )


def _budget_project_sheet_names(workbook) -> set[str]:
    if "BUDGET" not in workbook.sheetnames:
        return set()
    sheet = workbook["BUDGET"]
    project_names: set[str] = set()
    for row in range(10, 61):
        for col in ("B", "C", "D", "E"):
            value = sheet[f"{col}{row}"].value
            if isinstance(value, str) and value.strip().startswith("P-"):
                project_names.add(value.strip())
    return project_names


def parse_user_dimensions(sheet) -> list[UserDimensionRow]:
    rows = []
    for offset, row in enumerate(sheet[USER_RANGE], start=195):
        if offset == 195:
            continue
        person_slot = row[0].value
        userid = row[2].value
        if person_slot in (None, "") and userid in (None, ""):
            continue
        rows.append(
            UserDimensionRow(
                person_slot=str(person_slot).strip(),
                person_name=str(row[1].value).strip() if row[1].value not in (None, "") else "",
                userid=str(userid).strip(),
                gender=str(row[3].value).strip() if row[3].value not in (None, "") else None,
                role=str(row[4].value).strip() if row[4].value not in (None, "") else None,
                hourly_rate=_to_float(row[5].value),
                hourly_rate_overhead=_to_float(row[6].value),
                hour_divider=_to_float(row[7].value),
                contract_hours=_to_float(row[8].value),
                sandbox_pct=_to_float(row[9].value),
                source_sheet_title=sheet.title,
                source_row=offset,
            )
        )
    return rows


def parse_wp_dimensions(sheet) -> list[WPDimensionRow]:
    rows = []
    for offset, row in enumerate(sheet[WP_RANGE], start=210):
        if offset == 210:
            continue
        wp_code = row[0].value
        if wp_code in (None, ""):
            continue
        rows.append(
            WPDimensionRow(
                wp_code=str(wp_code).strip(),
                wp_shortname=str(row[1].value).strip() if row[1].value not in (None, "") else None,
                wp_long_name=str(row[2].value).strip() if row[2].value not in (None, "") else None,
            )
        )
    return rows


def _iter_month_columns() -> Iterable[int]:
    start = column_index_from_string(MONTH_START_COL)
    end = column_index_from_string(MONTH_END_COL)
    return range(start, end + 1)


def parse_project_sheet(sheet) -> list[ParsedProjectBlock]:
    month_header_row = _detect_month_header_row(sheet)
    month_headers = {
        col_idx: _excel_date_to_date(sheet.cell(row=month_header_row, column=col_idx).value)
        for col_idx in _iter_month_columns()
    }
    blocks: list[ParsedProjectBlock] = []
    for block_start in BLOCK_START_ROWS:
        person_slot = sheet[f"A{block_start}"].value
        person_name = sheet[f"B{block_start}"].value
        userid = sheet[f"B{block_start + 1}"].value
        role = sheet[f"C{block_start}"].value
        optional_attr = sheet[f"D{block_start}"].value
        summary_row = block_start + 8
        annual_summary = {
            f"annual_summary_{offset}": _to_float(sheet.cell(row=block_start, column=col_idx).value)
            for offset, col_idx in enumerate(range(column_index_from_string("F"), column_index_from_string("J") + 1), start=1)
        }
        summary_monthly_values = []
        has_summary_values = False
        for col_idx in _iter_month_columns():
            cell = sheet.cell(row=summary_row, column=col_idx)
            value = _to_float(cell.value) if cell.value not in ("", None) else None
            has_summary_values = has_summary_values or value is not None
            summary_monthly_values.append((month_headers[col_idx], cell.coordinate, value))
        if person_slot in (None, "") and userid in (None, "") and not has_summary_values:
            continue
        for row_idx in range(block_start, block_start + 8):
            wp_label_raw = sheet[f"E{row_idx}"].value
            monthly_values = []
            has_month_values = False
            for col_idx in _iter_month_columns():
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = _to_float(cell.value) if cell.value not in ("", None) else None
                has_month_values = has_month_values or value is not None
                monthly_values.append((month_headers[col_idx], cell.coordinate, value))
            if wp_label_raw in (None, "") and not has_month_values:
                continue
            blocks.append(
                ParsedProjectBlock(
                    sheet_title=sheet.title,
                    block_start_row=block_start,
                    source_row=row_idx,
                    person_slot=str(person_slot).strip() if person_slot not in (None, "") else None,
                    userid=str(userid).strip() if userid not in (None, "") else None,
                    person_name=str(person_name).strip() if person_name not in (None, "") else None,
                    role=str(role).strip() if role not in (None, "") else None,
                    optional_attr=str(optional_attr).strip() if optional_attr not in (None, "") else None,
                    wp_label_raw=str(wp_label_raw).strip() if wp_label_raw not in (None, "") else None,
                    annual_summary=annual_summary,
                    monthly_values=monthly_values,
                    summary_monthly_values=summary_monthly_values,
                )
            )
    return blocks


def _detect_month_header_row(sheet) -> int:
    for row_idx in (MONTH_HEADER_ROW, 4):
        if all(_looks_like_excel_date(sheet.cell(row=row_idx, column=col_idx).value) for col_idx in _iter_month_columns()):
            return row_idx
    raise ValueError(f"Could not detect month header row for sheet '{sheet.title}'.")
